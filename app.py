import os
from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from twilio.rest import Client
from datetime import datetime
import pandas as pd

# Load .env file
load_dotenv()

# Twilio credentials
TWILIO_ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID')
TWILIO_AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN')
TWILIO_WHATSAPP_NUMBER = os.getenv('TWILIO_WHATSAPP_NUMBER')

EXCEL_FILE = 'uploaded_excels/fees_data.xlsx'

app = Flask(__name__)
app.secret_key = 'your-secret-key'

UPLOAD_FOLDER = 'uploaded_excels'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Initialize Excel file if not exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append([
        "Student ID", "Name", "Date of Birth", "Gender", "Date of Admission",
        "Standard", "Division", "Parents Mobile No.", "Parent Email ID",
        "Fee Payable", "Payment Date", "Payment Mode", "Transaction ID",
        "Fees per Installment", "Amount Paid", "Remaining Balance",
        "Category", "Due Date", "Installment", "Fees Status", "Remarks"
    ])
    wb.save(EXCEL_FILE)

# Send WhatsApp message
def send_whatsapp_message(name, std, phone, amount):
    if not phone.startswith("whatsapp:"):
        phone = "whatsapp:+91" + phone.strip()

    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    message = (f"✅ *Payment Received - Kunjeer Public School* ✅\n\n"
               f"Dear Parent,\n\nWe have received ₹{amount} for *{name}* (Class: {std}).\n"
               f"Thank you for your timely payment.\n\nRegards,\nKunjeer Public School 🏫")

    try:
        client.messages.create(
            body=message,
            from_=TWILIO_WHATSAPP_NUMBER,
            to=phone
        )
        print("✅ WhatsApp message sent successfully!")
    except Exception as e:
        print("⚠️ Failed to send WhatsApp message:", e)

@app.route('/', methods=['GET', 'POST'])
def Fees_form():
    if request.method == 'POST':
        data = [
            request.form.get('student_id'),
            request.form.get('student_name'),
            request.form.get('dob'),
            request.form.get('gender'),
            request.form.get('admission_date'),
            request.form.get('standard'),
            request.form.get('division'),
            request.form.get('parent_mobile'),
            request.form.get('parent_email'),
            request.form.get('fee_payable'),
            request.form.get('payment_date'),
            request.form.get('payment_mode'),
            request.form.get('transaction_id'),
            request.form.get('fees_per_installment'),
            request.form.get('amount_paid'),
            request.form.get('remaining_balance'),
            request.form.get('category'),
            request.form.get('due_date'),
            request.form.get('installment'),
            request.form.get('fees_status'),
            request.form.get('remarks')
        ]

        excel_path = session.get('excel_path')
        if not excel_path or not os.path.exists(excel_path):
            return '❌ Please upload an Excel file first before submitting the form.'

        wb = load_workbook(excel_path)
        ws = wb.active
        ws.append(data)
        wb.save(excel_path)

        # Send WhatsApp
        send_whatsapp_message(
            name=request.form.get('student_name'),
            std=request.form.get('standard'),
            phone=request.form.get('parent_mobile'),
            amount=request.form.get('amount_paid')
        )

        return redirect(url_for('thank_you'))

    return render_template('form.html')

@app.route('/thanks')
def thank_you():
    return '''
    <h1 style="text-align:center; color:green;">✅ Thank you for your payment!</h1>
    <p style="text-align:center;">Your WhatsApp message will be delivered shortly.</p>
    <p style="text-align:center;"><a href="/">Go back to form</a></p>
    '''
@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    if 'excel_file' not in request.files:
        return 'No file part in the request', 400

    file = request.files['excel_file']

    if file.filename == '':
        return 'No selected file', 400

    if file:
        filename = file.filename
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        session['excel_path'] = filepath  # ✅ Store file path
        return '''
        <h2>✅ File uploaded successfully!</h2>
        <p><a href="/">Go back to form</a></p>
        '''

    return 'Something went wrong', 500
    
@app.route('/view_excel')
def view_excel():
    excel_path = session.get('excel_path')
    if not excel_path or not os.path.exists(excel_path):
        return '❌ No Excel file uploaded yet.'

    try:
        df = pd.read_excel(excel_path)
        table_html = df.to_html(classes='table table-striped', index=False)

        return f'''
        <html>
        <head>
            <title>Excel Preview</title>
            <style>
                body {{ font-family: Arial; padding: 20px; background: #f9f9f9; }}
                h2 {{ text-align: center; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                a {{ display: inline-block; margin-top: 20px; text-decoration: none; }}
            </style>
        </head>
        <body>
            <h2>📊 Uploaded Excel Sheet Preview</h2>
            {table_html}
            <br><br><a href="/">⬅️ Back to Form</a>
        </body>
        </html>
        '''
    except Exception as e:
        return f"⚠️ Error reading Excel file: {str(e)}"


if __name__ == '__main__':
    print("✅ Flask server starting...")
    app.run(debug=True)
    print('http://127.0.0.1:5000')
